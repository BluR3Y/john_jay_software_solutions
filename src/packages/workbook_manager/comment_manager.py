import openpyxl
import os

import openpyxl.comments

class CommentManager:
    comment_cache = {}
    
    def __init__(self, read_file_path: str = None, sheet_names: list[str] = None) -> None:
        if read_file_path:
            if not os.path.exists(read_file_path):
                raise FileNotFoundError(f"Invalid file path provided to CommentManager: {read_file_path}")
            
            # Load the workbook
            wb = openpyxl.load_workbook(read_file_path)
            existing_comments = {name: {} for name in wb.sheetnames}
            # Commented for debugging purposes ----- Intended to collect existing comments from read workbook
            # for sheet_name in workbook.sheetnames:
            #     sheet = workbook[sheet_name]
            #     existing_comments[sheet_name] = {
            #         f"{cell.row}:{cell.column}": cell.comment for row in sheet.iter_rows() for cell in row if cell.comment
            #     }
            self.comment_cache = existing_comments
            
        elif sheet_names:
            self.comment_cache = {name: {} for name in sheet_names}
        
    # Add cell comments to the object's comment cache
    def append_comment(self, sheet: str, row: int, col: int, comment: str, account_headers) -> None:
        """ Add a comment to a specific cell in the cache. """
        # if sheet not in self.sheets:
        #     raise Exception(f"The sheet '{sheet}' does not exist in the workbook")
        
        if sheet not in self.comment_cache:
            self.comment_cache[sheet] = {}
            
        if (row, col) not in self.comment_cache[sheet]:
            self.comment_cache[sheet][(row, col)] = [comment]
        else:
            self.comment_cache[sheet][(row, col)].append(comment)
            
    def create_comments(self, write_file_path):
        """ Write all cached comments to an Excel file. """
        if not os.path.exists(write_file_path):
            raise FileNotFoundError(f"Invalid file path provided: {write_file_path}")
        
        if not self.comment_cache:
            print("No comments to add. Exiting CommentManager.")
            return
        
        workbook = openpyxl.load_workbook(write_file_path)
        for sheet_name, comments in self.comment_cache.items():
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"The sheet '{sheet_name}' does not exist in the workbook.")
            
            sheet_content = workbook[sheet_name]
            for (row, col), comment_texts in comments.items():
                cell = sheet_content.cell(row + 1, col + 1) # Plus 1 accounts for rows and cols being 1-based index
                comment_text = ""
                for index, text in enumerate(comment_texts, start=1):
                    comment_text += f"Comment#{index}: {text}\n"
                comment_obj = openpyxl.comments.Comment(
                    comment_text, # Text
                    "Developer",  # Author
                    150, # Height in pixels
                    300  # Height in pixels
                )
                cell.comment = comment_obj
        # Save the workbook with the comments
        workbook.save(write_file_path)
        
# Rename module to styling_manager.py
# Include highlighting methods