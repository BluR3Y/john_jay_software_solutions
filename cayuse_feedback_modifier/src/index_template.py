from dotenv import load_dotenv
# Load environment variables from .env file
load_dotenv("../env/.env.development")

import openpyxl.comments
import pandas as pd
import os
import json
import argparse
import datetime
import openpyxl

import configs.db_config as db_config
import configs.openai_config as openai_config
import sheets.members as members
import sheets.attachments as attachments
import sheets.awards as awards
import sheets.proposals as proposals
from methods import utils
from methods import logger

class FeedBackModifier:
    
    def __init__(self):
        # Read Excel file
        self.df = pd.read_excel(os.get('EXCEL_FILE_PATH'), sheet_name=None)
        # Will be used to store comments before creating them in the excel file
        self.comment_cache = dict()
        
        self.logs_path = (os.getenv('SAVE_PATH') or os.path.dirname(self.filepath) )+ 'cayuse_data_migration_logs.json'
        if os.path.exists(self.logs_path):
            with open(self.logs_path) as json_file:
                self.logs = json.load(json_file)
        else:
            self.logs = dict()

    # Save changes to all related resources
    def save_changes(self, as_copy = True, index=False):
        try:
            excel_file_path = os.getenv('EXCEL_FILE_PATH')
            file_save_path = os.getenv('SAVE_PATH') if as_copy and os.getenv('SAVE_PATH') else os.path.dirname(excel_file_path)
            file_name = f"{os.path.splitext(os.path.basename(excel_file_path))[0]}{f" - modified_copy_{datetime.datetime.today().strftime('%m-%d-%Y')}" if as_copy else ''}.xlsx"
            full_path = os.path.join(file_save_path, file_name)

            # Use ExcelWriter to write multiple sheets back into the Excel file
            with pd.ExcelWriter(full_path, engine='openpyxl', mode='w') as writer:
                for sheet_name, df_sheet in self.df.items():
                    # Leave index as False if you don't want to save the index column from the DataFrame
                    df_sheet.to_excel(writer, sheet_name=sheet_name, index=index)

            # Load the workbook
            workbook = openpyxl.load_workbook(full_path)
            for sheet_name in self.comment_cache:
                if sheet_name in workbook.sheetnames:
                    sheet_content = workbook[sheet_name]
                    for cell_position in self.comment_cache[sheet_name]:
                        row, col = cell_position.split(':')
                        cell = sheet_content.cell(int(row) + 1, int(col) + 1)   # Plus 1 accounts for rows and columns being 1-based index
                        comment = openpyxl.comments.Comment(self.comment_cache[sheet_name][cell_position], "Developer")
                        cell.comment = comment
                        comment.height = 150 # Height in pixels
                        comment.width = 300 # Width in pixels
                else:
                    raise Exception(f"The sheet with the name '{sheet_name}' does not exist in the workbook")
            # Save the workbook with comments
            workbook.save(full_path)

            # Save the changes made to the logger
            self.save_logs()
        except Exception as e:
            print(f"Error saving Excel file: {e}")

    # Add cell comments to the object's comment cache
    def append_cell_comment(self, sheet, row, col, comment):
        if sheet not in self.comment_cache:
            self.comment_cache[sheet] = { f"{row}:{col}": comment }
        else:
            self.comment_cache[sheet].update({ f"{row}:{col}": comment })
        
# Database related methods
FeedBackModifier.execute_query = db_config.execute_query

# Utilities related methods
FeedBackModifier.grant_abondant_record_field_modifier = utils.grant_abondant_record_field_modifier

# Logger related methods
FeedBackModifier.save_logs = logger.save_logs
FeedBackModifier.append_logs = logger.append_logs

# Members sheet related methods
FeedBackModifier.modify_username = members.modify_entries

# Attachments sheet related methods
FeedBackModifier.verify_entries = attachments.verify_entries
FeedBackModifier.get_missing_project_attachments = attachments.missing_project_attachments
FeedBackModifier.get_project_info = attachments.retrieve_project_info

# Awards sheet related methods
FeedBackModifier.populate_db_award_discipline = awards.populate_db_discipline
FeedBackModifier.populate_template_award_discipline = awards.populate_template_discipline
FeedBackModifier.populate_template_award_department = awards.populate_template_department

# Proposal sheet related methods
FeedBackModifier.populate_template_proposal_discipline = proposals.populate_template_discipline
FeedBackModifier.populate_template_proposal_department = proposals.populate_template_department

# Testing the Modifier class:
if __name__ == "__main__":
    ## Create a new class instance
    my_instance = FeedBackModifier()

    # Initialize the argument parser
    parser = argparse.ArgumentParser(description="Parser handles command-line flags.")

    # Add optional flags and arguments
    # parser.add_argument('--save', '-s', type=str, help='Save changes to original file or as a copy')
    parser.add_argument('--methods', '-m', nargs='+', help='List of methods to call.')

    # Parse the arguments
    args = parser.parse_args()

    if args.methods:
        for method in args.methods:
            if hasattr(my_instance, method) and callable(getattr(my_instance, method)):
                getattr(my_instance, method)()
            else:
                raise Exception(f"The method {method} does not exist")
        # Save document changes
        my_instance.save_changes()