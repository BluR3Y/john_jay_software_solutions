import os
import pandas as pd
import openpyxl

from packages.workbook_manager import WorkbookManager
from modules.utils import request_user_selection, get_valid_filename

class ReportGenerator:
    process_name = "Report Generator"
    
    def __init__(self, save_path):
        self.save_path = save_path
        self.file_name = "generated_report"
        
    def __enter__(self):
        self.report_workbook = WorkbookManager()
        self.report_workbook.create_sheet("report_meta_data", ["sheet_name", "table", "record_identifier"])
        
        return self
    
    def __exit__(self, exc_type, exc_value, traceback):
        if self.get_num_reports() == 0:
            print("No reports were generated. Exiting Report Generator.")
            return
        
        save_path = {
            "dir_path": self.save_path,
            "file_name": self.file_name
        }
        
        if os.path.exists(os.path.join(save_path["dir_path"], f"{save_path['file_name']}.xlsx")):
            user_action = request_user_selection(f"A report currently exists with the name '{save_path['file_name']}'. Choose an action to take: ", ["Choose new file name", "Append to existing file", "Overwrite file"])
            if user_action == "Choose new file name":
                save_path['file_name'] = get_valid_filename()
            elif user_action == "Append to existing file":
                existing_workbook = WorkbookManager(os.path.join(save_path["dir_path"], f"{save_path['file_name']}.xlsx"))
                
                for sheet_name, sheet_content in self.report_workbook.df.items():
                    if sheet_name == "report_meta_data":
                        continue
                    
                    existing_workbook.create_sheet(sheet_name, sheet_content.to_dict(orient='records'))
                    sheet_meta_data = self.report_workbook.get_entries("report_meta_data", {"sheet_name": sheet_name})
                    existing_workbook.append_row("report_meta_data", {
                        "sheet_name": sheet_name,
                        "table": sheet_meta_data.get('table'),
                        "record_identifier": sheet_meta_data.get('record_identifier')
                    })
                self.report_workbook = existing_workbook

        try:
            self.report_workbook.save_changes(os.path.join(save_path["dir_path"], f"{save_path['file_name']}.xlsx"))
            print("Report successfully generated.")
        except Exception as err:
            print(f"Error occured while saving report: {err}")
    
    def append_report(self, report_name: str, table: str, record_identifier: str, report_data: list[dict]):
        self.report_workbook.create_sheet(report_name, report_data)
        self.report_workbook.append_row("report_meta_data", {
            "sheet_name": report_name,
            "table": table,
            "record_identifier": record_identifier
        })
        
    def get_num_reports(self) -> int:
        return len([item for item in self.report_workbook.df.keys() if item != "report_meta_data"])