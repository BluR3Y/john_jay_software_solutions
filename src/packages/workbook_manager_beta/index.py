import pandas as pd
from pathlib import Path
import openpyxl

from ..log_manager import LogManager
from . import SheetManager
from modules.utils import request_user_confirmation

class WorkbookManager:
    def __init__(self, read_file_path: str = None, write_file_path: str = None):
        """
        Initialize the Workbook Manager.

        Parameters:
            - read_file_path: Path to an existing Excel file to load.
            - write_file_path: Path to save changes when exiting the context manager.
        """
        if read_file_path and not Path(read_file_path).exists():
            raise ValueError(f"Workbook file does not exist at path: {read_file_path}")

        self.read_file_path = read_file_path
        self.write_file_path = write_file_path
        self.sheets = {}

    def __len__(self):
        return len(self.sheets)
    
    def __iter__(self):
        for sheet_name in self.df.keys():
            yield sheet_name

    def __getitem__(self, key):
        return self.sheets.get(key)

    def __enter__(self):
        """Context entry: load workbook data and initialize supporting managers."""
        if self.read_file_path:
            # Read the contents of the workbook
            excel_obj = pd.ExcelFile(self.read_file_path)

            for sheet_name in excel_obj.sheet_names:
                self.sheets[sheet_name] = SheetManager(excel_obj.parse(sheet_name))

            read_path_obj = Path(self.read_file_path)
            log_file_path = read_path_obj.parent / f"{read_path_obj.stem}_workbook_logs.json"
            self.log_manager = LogManager(log_file_path).__enter__()
        return self
    
    def __exit__(self, exc_type, exc_value, traceback):
        """Context exit: save workbook data and close log manager."""
        if not self.log_manager.runtime_date_time in self.log_manager.get_runtime_dates():
            print("No changes detected. Exiting Workbook Manager.")
            return

        if not self.read_file_path and not self.write_file_path:
            print("Workbook Manager is missing path to save file.")
            return
        
        if self.write_file_path:
            self._save_data()
        elif self.read_file_path:
            print("Workbook is missing write filepath.")
            if not self.set_write_path(self.read_file_path):
                return
            self._save_data()
            self.log_manager.__exit__(exc_type, exc_value, traceback)

    def set_write_path(self, write_file_path: str) -> bool:
        if write_file_path == '':
            raise ValueError(f"Empty string provided as write path.")
        
        # Confirm overwrite if same as original path
        if self.read_file_path and self.read_file_path == write_file_path:
            user_confirmation = request_user_confirmation(f"Are you sure you want to overwrite '{write_file_path}': ")
            if not user_confirmation:
                print("Workbook overwrite aborted.")
                return False
            
        self.write_file_path = write_file_path
        return True

    def _save_data(self):
        """
        Internal method to save workbook data to disk.
        """
        if not self.write_file_path:
            raise ValueError("Missing file write path")
        
        try:
            # Load existing workbook if available, otherwise create new one
            wb = None
            if self.read_file_path:
                wb = openpyxl.load_workbook(self.read_file_path)
            else:
                wb = openpyxl.Workbook()
                # Remove default "Sheet" only if its truly empty
                default_sheet = wb.active
                if default_sheet.title == "Sheet" and default_sheet.max_row == 1 and default_sheet.max_column == 1 and default_sheet["A1"].value is None:
                    wb.remove(default_sheet)

            for sheet_name, df_sheet in self.sheets.items():
                if sheet_name not in wb.sheetnames:
                    created_sheet = wb.create_sheet(sheet_name)
                    created_sheet.append(list(df_sheet.df.columns))
                
                df_sheet._append_data(wb[sheet_name])
            wb.save(self.write_file_path)    
        except Exception as err:
            raise Exception(f"Error occured while attempting to save Workbook data: {err}")
        
    def create_sheet(self, sheet_name: str, sheet_columns: list):
        """
        Add a new sheet to the workbook.

        Parameters:
        - sheet_name: Name of the new sheet.
        - sheet_columns: List of column headers or populated row dictionaries.
        """
        if sheet_name in self.sheets:
            raise ValueError(f"A sheet with the name '{sheet_name}' already exists in the workbook.")
        self.sheets[sheet_name] = SheetManager(sheet_columns)

    def get_sheet(self, sheet_name: str) -> SheetManager:
        """
        Retrieve a sheet as a dictionary. Optionally format and select specific columns.

        Parameters:
        - sheet_name: Sheet to fetch.
        """
        if sheet_name not in self.sheets:
            return None
        
        return self.sheets[sheet_name]
    
    def delete_sheet(self, sheet_name: str):
        if sheet_name not in self.sheets:
            return None
        
        del self.sheets[sheet_name]