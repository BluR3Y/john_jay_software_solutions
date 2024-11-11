import openpyxl
import os

class CommentManager:
    def __init__(self, read_file_path):
        if os.path.exists(read_file_path):
            existing_comments = dict()
            # Load the workbook
            workbook = openpyxl.load_workbook(read_file_path)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                existing_comments[sheet_name] = {
                    f"{cell.row}:{cell.column}": cell.comment for row in sheet.iter_rows() for cell in row if cell.comment
                }
            self.comment_cache = existing_comments
            self.sheets = workbook.sheetnames
            print(self.sheets)
        else:
            raise Exception("The comment manager was provided an invalid file path.")
        
    # Add cell comments to the object's comment cache
    def append_comment(self, sheet, row, col, comment):
        if sheet in self.sheets:
            if sheet not in self.comment_cache:
                self.comment_cache[sheet] = { f"{row + 1}:{col + 1}": comment } # Plus 1 accounts for rows and columns being 1-based index
            else:
                self.comment_cache[sheet].update({ f"{row + 1}:{col + 1}": comment })   # Plus 1 accounts for rows and columns being 1-based index
        else:
            raise Exception(f"The sheet '{sheet}' does not exist in the workbook")
        
    # Create the comments in the excel file that are stored in the object's cache
    def create_comments(self, write_file_path):
        # Load the workbook
        workbook = openpyxl.load_workbook(write_file_path)
        for sheet_name in self.comment_cache:
            if sheet_name in workbook.sheetnames:
                sheet_content = workbook[sheet_name]
                for cell_position in self.comment_cache[sheet_name]:
                    row, col = cell_position.split(':')
                    cell = sheet_content.cell(int(row), int(col))
                    comment = openpyxl.comments.Comment(self.comment_cache[sheet_name][cell_position], "Developer")
                    cell.comment = comment
                    comment.height = 150 # Height in pixels
                    comment.width = 300 # Width in pixels
            else:
                raise Exception(f"The sheet '{sheet_name}' does not exist in the workbook")
        # Save the workbook with the comments
        workbook.save(write_file_path)