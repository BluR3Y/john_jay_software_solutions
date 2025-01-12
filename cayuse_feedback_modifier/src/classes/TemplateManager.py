import os
import openpyxl.workbook
import pandas as pd
import openpyxl

from classes.LogManager.TemplateLogManager import TemplateLogManager

class TemplateManager:

    def __init__(self, read_file_path = None, log_file_path = None, create_sheets: dict = None):
        if read_file_path:
            # Store the file path in class instance
            self.read_file_path = read_file_path
            if os.path.exists(read_file_path):
                # Read the contents in the workbook
                self.df = pd.read_excel(read_file_path, sheet_name=None)
            else: 
                raise Exception("The Template manager was provided an invalid file path.")
        elif create_sheets:
            self.df = {sheet_name: pd.DataFrame({col_name: [] for col_name in props})
                       for sheet_name, props in create_sheets.items()}
        else:
            raise ValueError("Either read_file_path or create_sheets must be provided.")

        # Initialize an instance of the Logger
        self.log_manager = TemplateLogManager(log_file_path)
        
    def update_cell(self, process_name, sheet_name, row, col, new_val):
        if sheet_name in list(self.df.keys()):
            sheet_data_frame = self.df[sheet_name]
            cell_prev_value = sheet_data_frame.iloc[row][col]
            sheet_data_frame.loc[row, col] = new_val
            self.log_manager.append_log(
                process_name,
                sheet_name,
                row,
                col,
                cell_prev_value,
                new_val
            )
        else:
            raise Exception(f"The sheet with the name '{sheet_name}' does not exist in the workbook.")

    def get_entry(self, sheet_name: str, identifier: str, value: any, all: bool = False):
        """
        Retrieve rows from a specified sheet based on a column's value.

        Parameters:
        - sheet_name: Name of the sheet to search in.
        - identifier: Column name to filter by.
        - value: Value to search for in the specified column
        - all: If true, return all matching rows. If false, return the first match.

        Raises:
        - KeyError: If the sheet or column does not exist.
        """
        try:
            # Retrieve the sheet
            sheet_data_frame = self.df[sheet_name]
            # Filter rows based on the identifier and value
            matching_rows = sheet_data_frame[sheet_data_frame[identifier] == value]
            if all:
                # Return all matches as a DataFrame
                return matching_rows
            else:
                # Return first match or None
                return matching_rows.iloc[0] if not matching_rows.empty else None
        except KeyError as e:
            raise KeyError(f"{str(e)} not found in the workbook.")

    def save_changes(self, write_file_path: str, index=False):
        """
        Saves the data stored in the pandas dataframes by converting each dataframe into an excel sheet in the same workbook.
        Additionally, function also sets the 'sheet_state' property to that of the sheet in the workbook that was imported.
        
        # Parameters:
        - write_file_path: file path where the migration data will be stored.
        - index: Will determine if the index column from the DataFrame will persist in the stored excel sheet.
        """
        try:
            # Use ExcelWriter to write multiple sheets back into the Excel file
            with pd.ExcelWriter(write_file_path, engine='openpyxl', mode='w') as writer:
                for sheet_name, df_sheet in self.df.items():
                    df_sheet.to_excel(writer, sheet_name=sheet_name, index=index)
                    
            if self.read_file_path:
                # Load the feedback workbook
                feedback_wb = openpyxl.load_workbook(self.read_file_path)
                # Load the newly created workbook with the migration data
                migration_wb = openpyxl.load_workbook(write_file_path)
            
                for sheet in migration_wb.sheetnames:
                    if sheet in feedback_wb.sheetnames:
                        feedback_wb[sheet].sheet_state
                        migration_wb[sheet].sheet_state = feedback_wb[sheet].sheet_state
                    else:
                        migration_wb[sheet].sheet_state = "visible"
                        
                # Save the state changes made to the migration workbook
                migration_wb.save(write_file_path)
                    
            # Save logger changes
            self.log_manager.save_logs()
        except Exception as e:
            raise Exception(f"Error occured while attempting to save template data: {e}")