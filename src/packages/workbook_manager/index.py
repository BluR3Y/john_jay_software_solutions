import os
import pandas as pd
import pathlib
import warnings
import numpy as np
import openpyxl

from typing import Union
from modules.utils import request_user_confirmation

from . import PropertyManager
from ..log_manager import LogManager

class WorkbookManager:
    def __init__(self, read_file_path: str = None, write_file_path: str = None):
        """
        Initialize the WorkbookManager. Accepts either a read path or a dictionary to create sheets.

        Parameters:
        - read_file_path: Path to an existing Excel file to load.
        - write_file_path: Path to save changes when exiting the context manager.
        """
        if read_file_path:
            if not os.path.exists(read_file_path):
                raise ValueError(f"The WorkbookManager was provided an invalid file path: {read_file_path}")
        else:
            self.df = {}
            self.log_manager = None

        self.read_file_path = read_file_path
        self.write_file_path = write_file_path

    def __enter__(self):
        """Context entry: load workbook data and initialize supporting managers."""
        if self.read_file_path:
            # Read the contents of the workbook
            self.df = pd.read_excel(self.read_file_path, sheet_name=None)

            file_path_obj = pathlib.Path(self.read_file_path)
            log_file_path = os.path.join(file_path_obj.parent, f"{file_path_obj.stem}_workbook_logs.json")
            self.log_manager = LogManager(log_file_path).__enter__()

        # PropertyManager manages metadata/comments associated with sheets
        self.property_manager = PropertyManager(self.read_file_path, self.df.keys())
        return self
    
    def __exit__(self, exc_type, exc_value, traceback):
        """Context exit: save workbook data and close log manager."""
        if not self.read_file_path and not self.write_file_path:
            print("Workbook Manager is missing path to save file.")
            return

        if self.write_file_path:
            self._save_data(self.write_file_path)
        elif self.read_file_path:
            self._save_data(self.read_file_path)
            self.log_manager.__exit__(exc_type, exc_value, traceback)
        
        # Save workbook properties
        self.property_manager.apply_properties()
        
    def _save_data(self, write_path: str):
        """
        Internal method to save workbook data to disk.

        Parameters:
        - write_path: Destination path to save the Excel file.
        """
        if not write_path:
            raise ValueError("Missing file write path")
        
        # Confirm overwrite if same as original path
        if write_path == self.read_file_path:
            user_confirmation = request_user_confirmation(f"Are you sure you want to overwrite '{write_path}': ")
            if not user_confirmation:
                print("Workbook overwrite aborted.")
                return
        
        # Skip saving if workbook is empty (based on logs or data content)
        if self.read_file_path and not self.log_manager.get_runtime_logs(self.log_manager.runtime_date_time):
            print("Workbook is empty. Now exiting Workbook Manager.")
            return
        if not self.read_file_path and all([sheet_content.empty for sheet_content in self.df.values()]):
            print("Workbook is empty. Now exiting Workbook Manager.")
            return
        
        try:
            # Load existing workbook if available, otherwise create new one
            wb = openpyxl.load_workbook(self.read_file_path) if self.read_file_path else openpyxl.Workbook()
            for sheet_name, df_sheet in self.df.items():
                if sheet_name not in wb.sheetnames:
                    created_sheet = wb.create_sheet(sheet_name)
                    created_sheet.append(list(df_sheet.columns))
                sheet = wb[sheet_name]
                for r_idx, row in df_sheet.iterrows():
                    for c_idx, val in enumerate(row):
                        sheet.cell(row=r_idx + 2, column=c_idx + 1, value=val)
            wb.save(write_path)
        except Exception as err:
            raise Exception(f"Error occured while attempting to save Workbook data: {err}")    

    @staticmethod
    def formatDF(df: Union[pd.DataFrame, pd.Series]) -> Union[pd.DataFrame, pd.Series]:
        """
        Replace NaT and NaN with None, and convert datetime columns/values to native Python datetime.

        Parameters:
        - df: A pandas DataFrame or Series to format.
        """
        df = df.replace({pd.NaT: None, np.nan: None})

        if isinstance(df, pd.Series):
            # Convert datetime Series to Python datetimes
            if pd.api.types.is_datetime64_any_dtype(df):
                df = df.apply(lambda x: x.to_pydatetime() if pd.notnull(x) else None)
            return df
        
        for col in df.select_dtypes(include=["datetime64[ns]"]):
            df[col] = df[col].apply(lambda x: x.to_pydatetime() if pd.notnull(x) else None)
        return df

    def get_sheet(self, sheet_name: str, cols: list[str] = None, format = False, orient:str = None):
        """
        Retrieve a sheet as a dictionary. Optionally format and select specific columns.

        Parameters:
        - sheet_name: Sheet to fetch.
        - cols: Subset of columns to retrieve.
        - format: Apply formatDF for cleanup.
        - orient: Return format for dict (e.g., 'records', 'index').
        """
        if sheet_name not in self.df:
            return None
        
        cols = cols or []
        sheet_df = self.df[sheet_name][cols] if cols else self.df[sheet_name]
        sheet_df = sheet_df.copy()

        if format:
            sheet_df = self.formatDF(sheet_df)
        
        return sheet_df.to_dict(orient) if orient else sheet_df.to_dict()

    def create_sheet(self, sheet_name: str, sheet_columns: list):
        """
        Add a new sheet to the workbook.

        Parameters:
        - sheet_name: Name of the new sheet.
        - sheet_columns: List of column headers or populated row dictionaries.
        """
        if sheet_name in self.df.keys():
            raise ValueError(f"A sheet with the name '{sheet_name}' already exists in the workbook.")
        
        sheet_data_frame = None
        populated_data = all(isinstance(item, dict) for item in sheet_columns)
        if populated_data:
            sheet_data_frame = pd.DataFrame(sheet_columns)
        else:
            sheet_data_frame = pd.DataFrame(columns=sheet_columns)
            
        self.df[sheet_name] = sheet_data_frame
        self.property_manager.property_store[sheet_name] = {}
    
    def series_indices(self, sheet: str, obj: pd.Series):
        """
        Find row indices where the Series exactly matches a row in the sheet.
        """
        
        if sheet not in self.df:
            raise ValueError(f"The sheet '{sheet}' does not exist in the workbook.")
        
        sheet_df = self.df[sheet]

        if not isinstance(obj, pd.Series):
            raise TypeError("obj must be a pandas Series")

        # Ensure obj has the same columns as the DataFrame for valid comparison
        if not all(col in sheet_df.columns for col in obj.index):
            raise ValueError("Series contains columns that do not exist in the sheet.")

        matching_indices = sheet_df.index[sheet_df.apply(lambda row: row.equals(obj), axis=1)]
        return matching_indices.tolist()

    def update_cell(self, process: str, sheet: str, row: Union[int, pd.Series], properties: dict) -> pd.Series:
        """
        Update values in a given row of a sheet and log the changes.

        Parameters:
        - process: The process identifier for logging.
        - sheet: Target sheet name.
        - row: Target row (by index or Series match).
        - properties: Dictionary of updates to apply.
        """
        if sheet not in self.df:
            raise ValueError(f"Sheet '{sheet}' does not exist")
        sheet_df = self.df[sheet]

        # Determine the row index
        row_index = None
        if isinstance(row, pd.Series):
            matches = self.series_indices(sheet, row)
            if len(matches) != 1:
                raise ValueError("Series did not uniquely match a single row.")
            row_index = matches[0]
        elif isinstance(row, int):
            if not (0 <= row < sheet_df.shape[0]):
                raise ValueError("Row index out of bounds.")
            row_index = row
        else:
            raise TypeError("Row must be an int or pandas Series.")
        
        current = sheet_df.iloc[row_index]
        changing = {
            key: val for key, val in properties.items()
            if key in current.index and current[key] != val
        }

        # Warn on invalid property keys
        unknown_keys = set(properties.keys()) - set(current.index)
        if unknown_keys:
            warnings.warn(f"Ignored unknown column(s): {unknown_keys}")

        if not changing:
            return current
        
        # Apply changes
        for key, val in changing.items():
            self.df[sheet].at[row_index, key] = val

        # Log the update
        if self.log_manager:
            self.log_manager.append_runtime_log(process, "update", sheet, row_index, changing)

        return self.df[sheet].iloc[row_index]
        
    def append_row(self, process:str, sheet: str, props: dict):
        """
        Append a new row to the sheet.

        Parameters:
        - sheet: Target sheet name.
        - props: Dictionary of column values.
        """
        # Create a new DataFrame
        new_row = pd.DataFrame({key: [value] for key, value in props.items()})
        # Append using pd.concat
        self.df[sheet] = pd.concat([self.df[sheet], new_row], ignore_index=True)

        # Log the changes
        if self.log_manager:
            self.log_manager.append_runtime_log(process, "create", sheet, self.df[sheet].shape[0] - 1, props)

    def delete_row(self, process: str, sheet: str, row: Union[int, pd.Series]) -> None:
        """
        Deletes a row from the specified sheet in the workbook.

        Parameters:
        - process (str): Identifier used for logging the deletion.
        - sheet (str): Sheet name from which the row will be removed.
        - row (Union[int, pd.Series]): Row index or row content to identify the row to delete.
        """
        if sheet not in self.df:
            raise ValueError(f"The sheet '{sheet}' does not exist in the workbook.")

        sheet_df = self.df[sheet]

        # Determine row index
        if isinstance(row, int):
            if not (0 <= row < len(sheet_df)):
                raise ValueError(f"Row index {row} out of bounds.")
            row_index = row
        elif isinstance(row, pd.Series):
            matches = self.series_indices(sheet, row)
            if len(matches) != 1:
                raise ValueError("Series did not uniquely match a single row.")
            row_index = matches[0]
        else:
            raise TypeError("Row must be either an integer index or a pandas Series.")

        # Capture data before deletion for logging
        row_data = sheet_df.iloc[row_index].to_dict()

        # Drop the row and reset index
        self.df[sheet] = sheet_df.drop(index=row_index).reset_index(drop=True)

        # Log the deletion
        if self.log_manager:
            self.log_manager.append_runtime_log(process, "delete", sheet, row_index, row_data)

        
    def row_follows_condition(self, row, conditions: dict) -> bool:
        """Return True if all condition key-value pairs match the row."""
        return all(row.get(k) == v for k, v in conditions.items())
    
    def find(self, sheet_name: str, conditions: dict, return_one: bool = False, to_dict: str = None):
        """
        Retrieve rows matching specific conditions in a sheet.

        Parameters:
        - sheet_name: Target sheet.
        - conditions: Column-value pairs to match.
        - return_one: Return only the first match if True.
        - to_dict: If specified, return result as a dict with given orientation.
        """
        
        if sheet_name not in self.df:
            raise ValueError(f"The sheet '{sheet_name}' was not found in the workbook.")
        
        sheet_data_frame = self.df[sheet_name]
        
        # Ensure that conditions refer to existing columns
        invalid_columns = [col for col in conditions if col not in sheet_data_frame.columns]
        if invalid_columns:
            raise KeyError(f"Invalid column names in conditions: {invalid_columns}")
        
        # Apply filtering
        mask = pd.Series(True, index=sheet_data_frame.index)
        for column, value in conditions.items():
            mask &= (sheet_data_frame[column] == value)
        
        filtered_rows = sheet_data_frame[mask]
        
        if filtered_rows.empty:
            return filtered_rows
        
        formatted = self.formatDF(filtered_rows)
        
        if to_dict:
            results = formatted.to_dict(orient=to_dict)
            return results[0] if return_one else results
        return formatted.iloc[0] if return_one else formatted
    
    # def revert_transaction(self, )

    @staticmethod
    def values_equal(a, b):
        """
        Compare two values for equality, accounting for types like NaN, datetime, and strings.
        """
        if pd.isnull(a) and pd.isnull(b):
            return True
        if isinstance(a, (float, int)) and isinstance(b, str):
            try:
                return a == float(b)
            except ValueError:
                return False
        if isinstance(a, str) and isinstance(b, (float, int)):
            try:
                return float(a) == b
            except ValueError:
                return False
        if isinstance(a, pd.Timestamp):
            try:
                b_dt = pd.to_datetime(b)
                return a == b_dt
            except Exception:
                return False
        if isinstance(b, pd.Timestamp):
            try:
                a_dt = pd.to_datetime(a)
                return b == a_dt
            except Exception:
                return False
        return a == b

    @staticmethod
    def find_closest_row(base: dict, df: pd.DataFrame, threshold: float = 0.75) -> Union[pd.Series, None]:
        """
        Return the row in the DataFrame that most closely matches the given base dictionary.

        Parameters:
        - base: Dictionary of column-value pairs to compare against.
        - df: Target DataFrame to search.
        - threshold: Minimum match ratio required.
        """
        if df.empty:
            return None

        def num_matches(row: pd.Series) -> int:
            return sum(base.get(key) == row.get(key) for key in base)

        match_counts = {idx: num_matches(row) for idx, row in df.iterrows()}

        if not match_counts:
            return None

        best_idx = max(match_counts, key=match_counts.get)
        max_matches = match_counts[best_idx]
        total_possible = len(base)

        if total_possible == 0 or (max_matches / total_possible) < threshold:
            print(max_matches, total_possible, max_matches/total_possible)
            return None

        return df.loc[best_idx]