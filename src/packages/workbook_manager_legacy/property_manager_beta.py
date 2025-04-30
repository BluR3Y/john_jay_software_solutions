import openpyxl
import os
from typing import Literal

import openpyxl.comments
import openpyxl.styles

class PropertyManager:
    property_store = {
        "meta_data": {
            "sheet_state": "hidden",
            "cell_props": {
                (5,5): ("error","Test Error")
            }
        },
        "my_sheet": {
            "sheet_state": "visible",
            "cell_props": {
                (6,6): ("notice", "Test Notice")
            }
        }
    }


    def __init__(self):
        pass

    def append_sheet_property(self, sheet: str, prop_name, prop_val):
        sheet_properties = self.property_store.setdefault(sheet, {})
        sheet_properties[prop_name] = prop_val
        self.property_store[sheet] = sheet_properties

    def append_comment(self, sheet: str, row: int, col: int, type: Literal["notice", "warning", "error"], comment: str):
        sheet_properties = self.property_store.setdefault(sheet, {})

        if (row, col) not in sheet_properties:
            # sheet_properties[(row, col)] = [(type, comm)]
            # Last Here
            pass

    def apply_properties(self, write_file_path: str):
        if not write_file_path:
            raise ValueError("Failed to provide Comment Manager with write file path.")

        if not self.property_store:
            print("No workbook properties to apply. Exiting Comment Manager.")
            return
        
        wb = openpyxl.load_workbook(write_file_path)
        for sheet_name, sheet_props in self.property_store.items():
            sheet_content = wb[sheet_name]
            for prop_name, prop_val in [(key, val) for key, val in sheet_props.items() if not isinstance(val, dict)]:
                sheet_content[prop_name] = prop_val

            cell_props = sheet_props.get("cell_props")
            if cell_props:
                for (row, col), comments in cell_props.items():
                    cell = sheet_content.cell(row + 1, col + 1) # Plus 1 accounts for rows/cols being 1-based index
                    comment_text = ""
                    type_count = {"error": 0, "warning": 0, "notice": 0}
                    for type, text in comments:                        
                        comment_text += f"{type.capitalize()}#{type_count[type]+1}: {text}\n"
                        type_count[type] += 1
                    
                    comment_obj = openpyxl.comments.Comment(
                        comment_text, # Text
                        "Developer", # Author
                        150,  # Height in pixels
                        300   # Width in pixels
                    )
                    cell.comment = comment_obj

                    # Set cell fill color based on the most severe type
                    if type_count["error"] > 0:
                        fill_color = "bb2124"  # Red for errors
                    elif type_count["warning"] > 0:
                        fill_color = "ffcc00"  # Yellow for warnings
                    else:
                        fill_color = "5bc0de"  # Blue for notices
                            
                    cell.fill = openpyxl.styles.PatternFill(
                        start_color=fill_color,
                        end_color=fill_color,
                        fill_type='solid'
                    )
        wb.save(write_file_path)