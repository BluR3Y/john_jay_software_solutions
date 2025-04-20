import os
import pandas as pd
import openpyxl
import pathlib
import warnings
import numpy as np
from openpyxl import load_workbook

from typing import Union

from . import PropertyManager
from ..log_manager import LogManager

class WorkbookManager:    
    def __init__(self, read_file_path: str = None, create_sheets: dict = None, write_file_path: str = None):
        if not write_file_path:
            warn_msg = "meaning that file will be overwritten with changes." if read_file_path else "meaning workbook content won't be saved."
            warnings.warn(f"Write file path was not provided, {warn_msg}")

        if read_file_path:
            if not os.path.exists(read_file_path):
                raise ValueError(f"The WorkbookManager was provided an invalid file path: {read_file_path}")
            
            # Store the file path
            self.read_file_path = read_file_path
            
            # Read the contents of the workbook
            self.df = pd.read_excel(read_file_path, sheet_name=None)
        elif create_sheets:
            self.df = {sheet_name: pd.DataFrame(sheet_rows) for sheet_name, sheet_rows in create_sheets.items()}
        else:
            raise ValueError(f"Failed to pass workbook file path or sheet properties to WorkbookManager.")
        self.write_file_path = write_file_path
        
    def __enter__(self):
        if self.read_file_path:
            file_path_obj = pathlib.Path(self.read_file_path)
            log_file_path = os.path.join(file_path_obj.parent, f"{file_path_obj.stem}_workbook_logs.json")
            self.log_manager = LogManager(log_file_path).__enter__()

        # Initialize an instance of the Comment Manager
        self.property_manager = PropertyManager(self.read_file_path, self.df.keys())

        return self
    
    def __exit__(self, exc_type, exc_value, traceback):
        self._save_changes()
        if self.log_manager and self.write_file_path is None:
            self.log_manager.__exit__(exc_type, exc_value, traceback)

    # def _save_changes(self):
    #     if self.read_file_path:
    #         if not self.log_manager.get_runtime_logs(self.log_manager.runtime_date_time):
    #             print("Workbook is empty. Now exiting Workbook Manager.")
    #             return
            
    #         wb = load_workbook(self.read_file_path)
    #         for sheet_name, df_sheet in self.df.items():
    #             if sheet_name not in wb.sheetnames:
    #                 wb.create_sheet(sheet_name)
    #             sheet = wb[sheet_name]
    #             for r_idx, row in df_sheet.iterrows():
    #                 for c_idx, val in enumerate(row):
    #                     sheet.cell(row=r_idx + 2, column=c_idx + 1, value=val)
    #         wb.save(self.write_file_path or self.read_file_path)
    #     elif self.write_file_path:
    #         if all([sheet_content.empty for sheet_content in self.df.values()]):
    #             print("Workbook is empty. Now exiting Workbook Manager.")
    #             return
            
    #         with pd.ExcelWriter(self.write_file_path, engine='openpyxl', mode='w') as writer:
    #             for sheet_name, df_sheet in self.df.items():
    #                 df_sheet.to_excel(writer, sheet_name=sheet_name)
    #     else:
    #         return
    #     # Save Workbook Properties
    #     self.property_manager.apply_changes(self.write_file_path or self.read_file_path)

    def _save_changes(self):
        write_path = self.write_file_path or self.read_file_path
        if not write_path:
            print("Workbook Manager is missing path to save file")
            return
        
        if self.read_file_path and not self.log_manager.get_runtime_logs(self.log_manager.runtime_date_time):
            print("Workbook is empty. Now exiting Workbook Manager.")
            return
        
        if self.write_file_path and all([sheet_content.empty for sheet_content in self.df.values()]):
            print("Workbook is empty. Now exiting Workbook Manager.")
            return

        try:
            # Use ExcelWriter to write multiple sheets back into the Excel file
            with pd.ExcelWriter(write_path, engine='openpyxl', mode='w') as writer:
                for sheet_name, df_sheet in self.df.items():
                    df_sheet.to_excel(writer, sheet_name=sheet_name)
                    
            # Save LogManager changes
            if self.log_manager:
                self.log_manager.__exit__(None, None, None)
                
            # Save Workbook Properties
            self.property_manager.apply_changes(write_path)
        except Exception as err:
            raise Exception(f"Error occured while attempting to save Workbook data: {err}")            

    @staticmethod
    def formatDF(df: Union[pd.DataFrame, pd.Series]) -> Union[pd.DataFrame, pd.Series]:
        """
        Replace NaT and NaN with None, and convert datetime columns or values to Python datetime objects.
        Works for both Series and DataFrame inputs.
        """
        df = df.replace({pd.NaT: None, np.nan: None})

        if isinstance(df, pd.Series):
            # Convert datetime Series to Python datetimes
            if pd.api.types.is_datetime64_any_dtype(df):
                df = df.apply(lambda x: x.to_pydatetime() if pd.notnull(x) else None)
            return df
        
        for col in df.select_dtypes(include=["datetime64[ns]"]):
            df[col] = df[col].apply(lambda x: x.to_pydatetime() if pd.notnull(x) else None)
        return df

    def get_sheet(self, sheet_name: str, cols: list[str] = None, format = False, orient:str = None):
        if sheet_name not in self.df:
            return None
        
        cols = cols or []
        sheet_df = self.df[sheet_name][cols] if cols else self.df[sheet_name]
        sheet_df = sheet_df.copy()

        if format:
            sheet_df = self.formatDF(sheet_df)
        
        return sheet_df.to_dict(orient) if orient else sheet_df.to_dict()

    def create_sheet(self, sheet_name: str, sheet_columns: list):
        if sheet_name in self.df.keys():
            raise ValueError(f"A sheet with the name '{sheet_name}' already exists in the workbook.")
        
        sheet_data_frame = None
        populated_data = all(isinstance(item, dict) for item in sheet_columns)
        if populated_data:
            sheet_data_frame = pd.DataFrame(sheet_columns)
        else:
            sheet_data_frame = pd.DataFrame(columns=sheet_columns)
            
        self.df[sheet_name] = sheet_data_frame
        self.property_manager.property_store[sheet_name] = {}
    
    def series_indices(self, sheet: str, obj: pd.Series):
        """Return a list of row indices where the Series matches exactly in the given sheet."""
        
        if sheet not in self.df:
            raise ValueError(f"The sheet '{sheet}' does not exist in the workbook.")
        
        sheet_df = self.df[sheet]

        if not isinstance(obj, pd.Series):
            raise TypeError("obj must be a pandas Series")

        # Ensure obj has the same columns as the DataFrame for valid comparison
        if not all(col in sheet_df.columns for col in obj.index):
            raise ValueError("Series contains columns that do not exist in the sheet.")

        matching_indices = sheet_df.index[sheet_df.apply(lambda row: row.equals(obj), axis=1)]
        return matching_indices.tolist()

    def update_cell(self, process: str, sheet: str, row: Union[int, pd.Series], properties: dict) -> pd.Series:
        if sheet not in self.df:
            raise ValueError(f"Sheet '{sheet}' does not exist")
        sheet_df = self.df[sheet]

        row_index = None
        if isinstance(row, pd.Series):
            matches = self.series_indices(sheet, row)
            if len(matches) != 1:
                raise ValueError("Series did not uniquely match a single row.")
            row_index = matches[0]
        elif isinstance(row, int):
            if not (0 < row < sheet_df.shape[0]):
                raise ValueError("Row index out of bounds.")
            row_index = row
        else:
            raise TypeError("Row must be an int or pandas Series.")
        
        current = sheet_df.iloc[row_index]
        changing = {
            key: val for key, val in properties.items()
            if key in current.index and current[key] != val
        }

        unknown_keys = set(properties.keys()) - set(current.index)
        if unknown_keys:
            warnings.warn(f"Ignored unknown column(s): {unknown_keys}")

        if not changing:
            return current
        
        for key, val in changing.items():
            self.df[sheet].at[row_index, key] = val

        if self.log_manager:
            self.log_manager.append_runtime_log(process, "update", sheet, row_index, changing)

        return self.df[sheet].iloc[row_index]
        
    def append_row(self, sheet: str, props: dict):
        # Create a new DataFrame
        new_row = pd.DataFrame({key: [value] for key, value in props.items()})
        # Append using pd.concat
        self.df[sheet] = pd.concat([self.df[sheet], new_row], ignore_index=True)
        
    def row_follows_condition(self, row, conditions: dict) -> bool:
        return all(row.get(k) == v for k, v in conditions.items())
    
    def find(self, sheet_name: str, conditions: dict, return_one: bool = False, to_dict: str = None):
        """
        Retrieve rows from a specified sheet in the workbook that match the given column-value conditions.

        Args:
            sheet_name (str): Name of the sheet to query.
            conditions (dict): Dictionary of {column: value} to filter the DataFrame.
            return_one (bool, optional): If True, return only the first matching row. Defaults to False.
            to_dict (str, optional): If provided, converts result to dictionary using the specified orientation (e.g., 'records').

        Returns:
            pd.DataFrame, dict, or None: Filtered result(s), format depending on `return_one` and `to_dict`.
        """
        
        if sheet_name not in self.df:
            raise ValueError(f"The sheet '{sheet_name}' was not found in the workbook.")
        
        sheet_data_frame = self.df[sheet_name]
        
        # Ensure that conditions refer to existing columns
        invalid_columns = [col for col in conditions if col not in sheet_data_frame.columns]
        if invalid_columns:
            raise KeyError(f"Invalid column names in conditions: {invalid_columns}")
        
        # Apply filtering
        mask = pd.Series(True, index=sheet_data_frame.index)
        for column, value in conditions.items():
            mask &= (sheet_data_frame[column] == value)
        
        filtered_rows = sheet_data_frame[mask]
        
        if filtered_rows.empty:
            return filtered_rows
        
        formatted = self.formatDF(filtered_rows)
        
        if to_dict:
            results = formatted.to_dict(orient=to_dict)
            return results[0] if return_one else results
        return formatted.iloc[0] if return_one else formatted

    @staticmethod
    def values_equal(a, b):
        if pd.isnull(a) and pd.isnull(b):
            return True
        if isinstance(a, (float, int)) and isinstance(b, str):
            try:
                return a == float(b)
            except ValueError:
                return False
        if isinstance(a, str) and isinstance(b, (float, int)):
            try:
                return float(a) == b
            except ValueError:
                return False
        if isinstance(a, pd.Timestamp):
            try:
                b_dt = pd.to_datetime(b)
                return a == b_dt
            except Exception:
                return False
        if isinstance(b, pd.Timestamp):
            try:
                a_dt = pd.to_datetime(a)
                return b == a_dt
            except Exception:
                return False
        return a == b

    @staticmethod
    def find_closest_row(base: dict, df: pd.DataFrame, threshold: float = 0.75) -> Union[pd.Series, None]:
        if df.empty:
            return None

        def num_matches(row: pd.Series) -> int:
            return sum(base.get(key) == row.get(key) for key in base)

        match_counts = {idx: num_matches(row) for idx, row in df.iterrows()}

        if not match_counts:
            return None

        best_idx = max(match_counts, key=match_counts.get)
        max_matches = match_counts[best_idx]
        total_possible = len(base)

        if total_possible == 0 or (max_matches / total_possible) < threshold:
            return None

        return df.loc[best_idx]