from __future__ import annotations

from typing import Any, Dict, Iterable, Mapping, Optional, Sequence
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

from .sheet_manager import SheetManager

from modules.script_logger import logger
log = logger.get_logger()

class WorkbookManager:
    """
    Orchestrates reading/writing Excel files and registering SheetManagers.

    Notes:
      - This class separates library logic from any CLI/user interaction.
      - Overwrite behavior is explicit via set_write_path(..., allow_overwrite=True).
      - A 'dirty' flag tracks mutations to decide whether saving is necessary.
    """
    def __init__(self, read_file_path: Optional[str] = None) -> None:
        self.read_file_path = read_file_path
        self.write_file_path: Optional[str] = None
        self._wb = None # openpyxl workbook (lazily loaded)
        self._sheets: Dict[str, SheetManager] = {}
        self._dirty: bool = False
    
    # --------------------------- Workbook I/O ---------------------------
    def open(self) -> None:
        if self.read_file_path:
            log.info("Opening workbook: %s", self.read_file_path)
            self._wb = load_workbook(self.read_file_path, data_only=True)
            self.load_existing_sheets()
        else:
            log.info("Creating in-memory workbook.")
            self._wb = Workbook()
    
    def close(self) -> None:
        self._wb = None

    def load_existing_sheets(self, sheet_names: Optional[Sequence[str]] = None, **read_excel_kwargs) -> None:
        """
        (Re)load sheets from self.read_file_path into SheetManagers.

        Args:
            sheet_names: Optional list of sheet names to load. If None, load all sheets.
            **read_excel_kwargs: Extra kwargs passed to pandas.read_excel (e.g., dtype, header).
        """
        if not self.read_file_path:
            raise ValueError("No read_file_path is set; cannot load existing sheets.")
        if sheet_names is None:
            book = pd.read_excel(self.read_file_path, sheet_name=None, **read_excel_kwargs)
        else:
            book = pd.read_excel(self.read_file_path, sheet_name=list(sheet_names), **read_excel_kwargs)
            if isinstance(book, pd.DataFrame):
                only = sheet_names[0] if sheet_names else "Sheet1"
                book = {only: book}
        self._sheets.clear()
        for name, df in book.items():
            self._sheets[str(name)] = SheetManager(str(name), df)
        self._dirty = False
    
    def set_write_path(self, path: str, *, allow_overwrite: bool = False) -> None:
        if not path:
            raise ValueError("Empty write path.")
        if self.read_file_path and path == self.read_file_path and not allow_overwrite:
            raise ValueError("Refusing to overwrite source file without allow_overwrite=True.")
        self.write_file_path = path

    # --------------------------- Sheet registration ---------------------------
    def register_sheet(self, name: str, df: pd.DataFrame) -> SheetManager:
        sm = SheetManager(name, df.copy())
        self._sheets[name] = sm
        self._dirty = True
        return sm
    
    def get_sheet(self, name: str) -> SheetManager:
        return self._sheets[name]
    
    def list_sheets(self) -> Sequence[str]:
        return list(self._sheets.keys())
    
    # --------------------------- Mutations ---------------------------
    def mark_dirty(self) -> None:
        self._dirty = True

    # --------------------------- Save ---------------------------
    def save(self) -> Optional[str]:
        """
        Save registered sheets into `write_file_path`. Creates/overwrites the file depending on set_write_path().
        Returns the path if a save occurred, else None.
        """
        if not self.write_file_path:
            log.info("No write path set; skipping save.")
            return None
        if not self._dirty and self._sheets:
            log.info("No changes to save; skipping save.")
            return None
        
        try:
            wb = Workbook()
            # Remove default sheet
            default = wb.active
            wb.remove(default)

            for name, sm in self._sheets.items():
                ws = wb.create_sheet(title=name)
                df = sm.get_df(format=True)
                # Write handler
                ws.append(list(df.columns))
                # Write rows
                for _, row in df.iterrows():
                    ws.append(list(row.values))
                
                # Apply annotations
                self._apply_annotations(ws, sm)

            wb.save(self.write_file_path)
            log.info("Saved workbook to %s", self.write_file_path)
            self._dirty = False
            return self.write_file_path
        except Exception as err:
            raise Exception("Failed to save workbook data.") from err
        
    def _apply_annotations(self, ws, sm: SheetManager) -> None:
        """
        Render recorded annotations as comments and cell background fills.
        """
        fills = {
            "info":  PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"),
            "warn":  PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
            "error": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
        }
        for (row, col, level, message) in sm.iter_annotations():
            # +2 offset: Excel is 1-based and we have a header row at row 1
            erow = row + 2
            ecol = col + 1
            cell = ws.cell(row=erow, column=ecol)
            if message:
                if cell.comment:
                    cell.comment = Comment(cell.comment.text + "\\n" + message, "wbm")
                else:
                    cell.comment = Comment(message, "wbm")
            if level in fills:
                cell.fill = fills[level]

    # --------------------------- Context manager ---------------------------
    def __enter__(self) -> "WorkbookManager":
        self.open()
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        try:
            if exc is None:
                self.save()
        finally:
            self.close()