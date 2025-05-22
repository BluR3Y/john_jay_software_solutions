import pandas as pd
import logging
from typing import Union, Literal
from openpyxl import worksheet
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import numpy as np
from openpyxl.cell import cell

from datetime import datetime, date, time
from decimal import Decimal

assignable_sheet_props = ["sheet_state","column_dimensions"]
assignable_cell_props = ["comment","fill","border","font","hyperlink","value","alignment","number_format"]

class SheetManager:
    # openpyxl sheet properties: Title, col/row height/width, lock row/col, tab color, hide/Unhide col/rows, lock cells, modify print settings, hide sheets, add header/footer, Set gridline visibility
    
    # Reference
    # assigned_sheet_props = {
    #     "sheet_state": "visible"
    # }

    # openpyxl cell properties: Value, Font, Fill (Background Color), Border, Alignment, Number Format, Comment, Hyperlink, Merge Cells, Unmerge Cells
    
    # Reference
    # assigned_cell_props = {
    #     (1,5): {
    #         "comment": Comment("Test Comment Browski", "Developer", 150, 300),
    #         "fill": PatternFill(start_color='bb2124',end_color='bb2124', fill_type='solid')
    #     }
    # }

    issue_types = {
        "notice": {
            "priority": 2,
            "color": "5bc0de"
        },
        "warning": {
            "priority": 1,
            "color": "ffcc00"
        },
        "error": {
            "priority": 0,
            "color": "bb2124"
        }
    }

    def __iter__(self):
        for index, row in self.df.iterrows():
            yield row.to_dict()

    def __getitem__(self, key) -> Union[pd.Series, pd.DataFrame]:
        if not isinstance(key, int) and not isinstance(key, list):
            raise ValueError("Invalid key type")
        return self.format_df(self.df.iloc[key])

    def __init__(self, sheet_data: Union[pd.DataFrame, list]):
        if isinstance(sheet_data, list):
            populated_data = all(isinstance(item, dict) for item in sheet_data)
            self.df = pd.DataFrame(sheet_data) if populated_data else pd.DataFrame(columns=sheet_data)
        elif isinstance(sheet_data, pd.DataFrame):
            self.df = sheet_data
        else:
            raise TypeError("Invalid sheet_data provided to SheetManager.")
        self.assigned_sheet_props = {}
        self.assigned_cell_props = {}

    def _append_data(self, dest_sheet: worksheet):
        # Apply assigned sheet properties
        for attr_name, attr_val in self.assigned_sheet_props.items():
            setattr(dest_sheet, attr_name, attr_val)

        # Identify empty columns (i.e., all values are NaN or None)
        empty_columns = set()
        for c_idx in range(self.df.shape[1]):
            if self.df.iloc[:, c_idx].isnull().all():
                empty_columns.add(c_idx)

        # Populate the sheet
        for r_idx, row in self.df.iterrows():
            for c_idx, val in enumerate(row):
                curr_cell = dest_sheet.cell(row=r_idx + 2, column=c_idx + 1)

                if isinstance(curr_cell, cell.MergedCell):
                    continue

                curr_cell.value = val

                cell_attrs = self.assigned_cell_props.get((r_idx, c_idx))
                if cell_attrs:
                    for attr_name, attr_val in cell_attrs.items():
                        setattr(curr_cell, attr_name, attr_val)
        
        # Hide columns with no values
        for c_idx in empty_columns:
            col_letter = get_column_letter(c_idx + 1)
            dest_sheet.column_dimensions[col_letter].hidden = True

    def series_indices(self, obj: Union[pd.Series, dict], treat_nan_equal: bool = False) -> list[int]:
        """
        Find row indices in the sheet where the given Series or dict exactly matches.

        Parameters:
            obj (Union[pd.Series, dict]): Object to match against rows.
            treat_nan_equal (bool): If True, treat NaNs as equal in comparisons.

        Returns:
            list[int]: Indices of rows that exactly match the given object.
        """
        # Normalize input to Series
        if isinstance(obj, pd.Series):
            obj_series = obj
        elif isinstance(obj, dict):
            obj_series = pd.Series(obj)
        else:
            raise TypeError("Object must be either a pandas Series or a dictionary.")
        
        # Validate column alignment
        if not all(col in self.df.columns for col in obj_series.index):
            raise ValueError("Series contains columns that do not exist in the sheet.")
        
        if self.df.empty:
            return []

        # Handle NaN-safe comparison if specified
        df_cmp = self.df[obj_series.index]
        obj_cmp = obj_series.values
        if treat_nan_equal:
            df_cmp = df_cmp.fillna("<<NA>>")
            obj_cmp = obj_series.fillna("<<NA>>").values

        # Find matching rows
        matching_mask = (df_cmp == obj_cmp).all(axis=1)
        return self.df.index[matching_mask].tolist()
    
    def update_cell(self, row: Union[int, pd.Series], properties: dict) -> pd.Series:
        """
        Update values in a given row of a sheet and log the changes.

        Parameters:
        - row: Target row (by index or Series match).
        - properties: Dictionary of updates to apply.
        """
        # Determine the row index
        row_index = None
        if isinstance(row, pd.Series):
            matches = self.series_indices(row)
            if len(matches) != 1:
                raise ValueError("Series did not uniquely match a single row.")
            row_index = matches[0]
        elif isinstance(row, int):
            if not (0 <= row < self.df.shape[0]):
                raise ValueError("Row index out of bounds")
            row_index = row
        else:
            raise TypeError("Row must be an int or pandas Series")
        
        current = self.df.iloc[row_index]
        changing = {
            key: val for key, val in properties.items()
            if key in current.index and current[key] != val
        }

        # Warn on invalid property keys
        unknown_keys = set(properties.keys()) - set(current.index)
        if unknown_keys:
            logging.warning(f"Ignored unknown column(s): {unknown_keys}")
        
        if not changing:
            return current
        
        # Apply changes
        for key, val in changing.items():
            self.df.at[row_index, key] = val

        return self.df.iloc[row_index]

    def append_row(self, props: dict) -> pd.Series:
        invalid_assigns = []
        valid_assigns = {}

        for key, val in props.items():
            if self.valid_assign(val):
                valid_assigns[key] = [val]
            else:
                invalid_assigns.append(f"{key} - {val}")
        
        if invalid_assigns:
            raise TypeError(f"Properties included invalid assigns:\n {'\n'.join(invalid_assigns)}")
        
        # Create a new DataFrame
        new_row = pd.DataFrame(valid_assigns)
        
        # Append using pd.concat
        self.df = pd.concat([self.df, new_row], ignore_index=True)
        return new_row.iloc[0]

    def delete_row(self, row: Union[int, pd.Series]) -> pd.Series:
        """
        Deletes a row from the specified sheet in the workbook.

        Parameters:
        - row (Union[int, pd.Series]): Row index or row content to identify the row to delete.
        """
        # Determine the row index
        row_index = None
        if isinstance(row, pd.Series):
            matches = self.series_indices(row)
            if len(matches) != 1:
                raise ValueError("Series did not uniquely match a single row.")
            row_index = matches[0]
        elif isinstance(row, int):
            if not (0 <= row < self.df.shape[0]):
                raise ValueError("Row index out of bounds")
            row_index = row
        else:
            raise TypeError("Row must be an int or pandas Series")
        
        # Capture data before deletion
        row_data = self.df.iloc[row_index]

        # Drop the row and reset index
        self.df = self.df.drop(index=row_index).reset_index(drop=True)
        return row_data
    
    def find(self, conditions: dict, return_one: bool = False, clean_copy=False):
        """
        Retrieve rows matching specific conditions in a sheet.

        Parameters:
        - conditions: Column-value pairs to match.
        - return_one: Return only the first match if True.
        - to_dict: If specified, return result as a dict with given orientation.
        """
        # Ensure that conditions refer to existing columns
        invalid_columns = [col for col in conditions if col not in self.df.columns]
        if invalid_columns:
            raise KeyError(f"Invalid column names in conditions: {invalid_columns}")
        
        # Apply filtering
        mask = pd.Series(True, index=self.df.index)
        for column, value in conditions.items():
            mask &= (self.df[column] == value)
        
        filtered_rows = self.df[mask]
        
        if filtered_rows.empty:
            return filtered_rows
        
        formatted = self.format_df(filtered_rows)
        
        if clean_copy:
            result = formatted.copy()
            return result.reset_index(drop=True)

        return formatted.iloc[0] if return_one else formatted
    
    def add_issue(self, row: int, col: Union[int, str], type: Literal["notice", "warning", "error"], message: str):
        """Alter cell properties to indicate warning state."""
        if isinstance(col, str):
            sheet_cols = self.df.columns
            col_index = next(index for index, value in enumerate(sheet_cols) if value == col)
            if col_index is None:
                raise ValueError(f"{col} is not a valid column name in the sheet.")
            col = col_index
        
        num_rows, num_cols = self.df.shape
        if not (0 <= row <= num_rows) or not (0 <= col < num_cols):
            raise ValueError("Index is out of bounds.")

        issue_props = self.issue_types[type]
        cell_key = (row, col)
        cell_props = self.assigned_cell_props.get(cell_key, {})
        comment_obj = cell_props.get("comment")

        # Initial comment and fill
        if not comment_obj:
            new_comment = Comment(f"Issues:\n* {type.capitalize()}#1 - {message}", "Developer", width=150, height=300)
            self.assigned_cell_props[cell_key] = {
                **cell_props,
                "comment": new_comment,
                "fill": PatternFill(start_color=issue_props["color"], end_color=issue_props["color"], fill_type='solid')
            }
            return

        # Extract and update issues from existing comment
        cell_messages = comment_obj.text.splitlines()
        try:
            issues_start_index = cell_messages.index("Issues:")
        except ValueError:
            issues_start_index = len(cell_messages)
            cell_messages.append("Issues:")

        issues_lines = cell_messages[issues_start_index + 1:]
        issues_by_type = {}
        min_priority = issue_props
        for issue in issues_lines:
            if issue.startswith("* "):
                try:
                    prefix, msg = issue[2:].split(" - ", 1)
                    issue_type = prefix.split("#")[0].lower()
                    issues_by_type.setdefault(issue_type, []).append(msg)
                    if self.issue_types[issue_type]["priority"] < min_priority["priority"]:
                        min_priority = self.issue_types[issue_type]
                except Exception:
                    continue

        # Append the new issue
        issues_by_type.setdefault(type, []).append(message)

        # Rebuild comment text
        formatted_issues = [
            f"* {itype.capitalize()}#{idx+1} - {msg}"
            for itype, messages in issues_by_type.items()
            for idx, msg in enumerate(messages)
        ]
        new_comment_text = "\n".join([*cell_messages[:issues_start_index + 1], *formatted_issues])
        self.assigned_cell_props[cell_key]["comment"].text = new_comment_text

        # Update fill based on highest priority
        self.assigned_cell_props[cell_key]["fill"] = PatternFill(
            start_color=min_priority['color'],
            end_color=min_priority['color'],
            fill_type='solid'
        )

    def get_df(self, cols: list[str] = [], format: bool = False) -> pd.DataFrame:
        sheet_df = self.df[cols] if len(cols) else self.df
        sheet_df = sheet_df.copy()

        if format:
            sheet_df = self.format_df(sheet_df)
        
        return sheet_df

    # def find_differences(self, target: "SheetManager", identifier_cols: list[str]) -> dict:
    #     """
    #     Compares rows between the current sheet and a target sheet using identifier columns.
    #     Returns a dictionary where keys are row indices from self.df, and values are either:
    #         - A dict of differing field values from the target (excluding identifier columns)
    #         - None, if no matching row was found in the target

    #     Notes:
    #         - Rows with multiple source matches are processed using closest-match logic.
    #         - Target matches are consumed greedily to avoid reuse in fuzzy matching.
    #         - NaN values are considered equal.
    #     """
        
    #     # Validate identifier columns
    #     if not identifier_cols:
    #         raise ValueError("Failed to provide identifier_cols")
        
    #     shared_cols = list(set(self.df.columns) & set(target.df.columns))
    #     if not shared_cols:
    #         raise KeyError("Sheets don't share any similar columns.")

    #     if not all(col in shared_cols for col in identifier_cols):
    #         raise ValueError("Invalid identifier columns.")

    #     source_df = self.get_df(format=True)

    #     # Extract unique identifier tuples
    #     unique_identifiers = {
    #         tuple(row[col] for col in identifier_cols)
    #         for _, row in source_df[identifier_cols].iterrows()
    #     }

    #     # Helper: NaN-safe value diffing
    #     def get_changes(base: dict, ref: dict) -> dict:
    #         def differs(a, b):
    #             if pd.isna(a) and pd.isna(b):
    #                 return False
    #             return a != b

    #         return {
    #             key: val for key, val in ref.items()
    #             if key not in identifier_cols and differs(base.get(key), val)
    #         }

    #     differences = {}

    #     for identifier in unique_identifiers:
    #         # Build dict for row lookup
    #         record_identifier = {col: identifier[i] for i, col in enumerate(identifier_cols)}
            
    #         source_matches_ref = self.find(record_identifier)
    #         target_matches_ref = target.find(record_identifier, clean_copy=True)

    #         # Create dict list of source matches
    #         source_matches = source_matches_ref.to_dict(orient='records')
    #         # Map local index to original DataFrame index
    #         source_mapping = {
    #             local_idx: global_idx
    #             for local_idx, (global_idx, _) in enumerate(source_matches_ref.iterrows())
    #         }

    #         if len(source_matches_ref) > 1:
    #             # If no target match at all, mark all source rows as unmatched
    #             if target_matches_ref is None or target_matches_ref.empty:
    #                 for global_idx in source_mapping.values():
    #                     differences[global_idx] = None
    #                 continue

    #             # Sort source rows by count of non-null fields (most complete first)
    #             sorted_local_indices = sorted(
    #                 range(len(source_matches)),
    #                 key=lambda idx: sum(val is not None for val in source_matches[idx].values()),
    #                 reverse=True
    #             )

    #             for local_idx in sorted_local_indices:
    #                 source_row = source_matches[local_idx]
    #                 global_idx = source_mapping[local_idx]

    #                 if not target_matches_ref.empty:
    #                     closest_idx = self.find_closest_row(source_row, target_matches_ref, threshold=0.15)
    #                     target_row = target_matches_ref.iloc[closest_idx].to_dict()

    #                     changes = get_changes(source_row, target_row)
    #                     if changes:
    #                         differences[global_idx] = changes

    #                     # Remove the used target match to prevent reuse
    #                     target_matches_ref.drop(index=target_matches_ref.index[closest_idx], inplace=True)
    #                 else:
    #                     differences[global_idx] = None

    #         else:
    #             # Single source match
    #             source_row = source_matches[0]
    #             global_idx = source_mapping[0]

    #             if target_matches_ref is None or target_matches_ref.empty:
    #                 differences[global_idx] = None
    #             elif len(target_matches_ref) > 1:
    #                 closest_idx = self.find_closest_row(source_row, target_matches_ref, threshold=0.15)
    #                 target_row = target_matches_ref.iloc[closest_idx].to_dict()
    #                 changes = get_changes(source_row, target_row)
    #                 if changes:
    #                     differences[global_idx] = changes
    #             else:
    #                 # Single target match
    #                 target_row = target_matches_ref.iloc[0].to_dict()
    #                 changes = get_changes(source_row, target_row)
    #                 if changes:
    #                     differences[global_idx] = changes

    #     return differences

    def find_differences(self, target: "SheetManager", identifier_cols: list[str], checking_cols: list[str] = None) -> dict:
        """
        Compares rows between the current sheet and a target sheet using identifier columns.
        Returns a dictionary where keys are row indices from self.df, and values are either:
            - A dict of differing field values from the target (excluding identifier columns)
            - None, if no matching row was found in the target

        Notes:
            - Rows with multiple source matches are processed using closest-match logic.
            - Target matches are consumed greedily to avoid reuse in fuzzy matching.
            - NaN values are considered equal.
        """
        
        # Validate identifier columns
        if not identifier_cols:
            raise ValueError("Failed to provide identifier_cols")
        
        shared_cols = list(set(self.df.columns) & set(target.df.columns))
        if not shared_cols:
            raise KeyError("Sheets don't share any similar columns.")

        if not all(col in shared_cols for col in identifier_cols):
            raise ValueError("Invalid identifier columns.")
        
        # Determine columns to check
        if checking_cols:
            if not all(col in shared_cols and col not in identifier_cols for col in checking_cols):
                raise ValueError("Invalid checking columns.")
        else:
            checking_cols = [col for col in shared_cols if col not in identifier_cols]

        source_df = self.get_df(format=True)

        # Extract unique identifier tuples
        unique_identifiers = {
            tuple(row[col] for col in identifier_cols)
            for _, row in source_df[identifier_cols].iterrows()
        }

        def get_changes(base: dict, ref: dict) -> dict:
            """Return a dict of changed fields from `ref`, limited to `checking_cols`."""
            def differs(a, b):
                if pd.isna(a) and pd.isna(b):
                    return False
                return a != b

            return {
                key: ref[key]
                for key in checking_cols
                if differs(base.get(key), ref.get(key))
            }

        differences = {}

        for identifier in unique_identifiers:
            # Build dict for row lookup
            record_identifier = {col: identifier[i] for i, col in enumerate(identifier_cols)}
            
            source_matches_ref = self.find(record_identifier)
            target_matches_ref = target.find(record_identifier, clean_copy=True)

            # Create dict list of source matches
            source_matches = source_matches_ref.to_dict(orient='records')
            # Map local index to original DataFrame index
            source_mapping = {
                local_idx: global_idx
                for local_idx, (global_idx, _) in enumerate(source_matches_ref.iterrows())
            }

            if len(source_matches_ref) > 1:
                # If no target match at all, mark all source rows as unmatched
                if target_matches_ref is None or target_matches_ref.empty:
                    for global_idx in source_mapping.values():
                        differences[global_idx] = None
                    continue

                # Sort source rows by count of non-null fields (most complete first)
                sorted_local_indices = sorted(
                    range(len(source_matches)),
                    key=lambda idx: sum(val is not None for val in source_matches[idx].values()),
                    reverse=True
                )

                for local_idx in sorted_local_indices:
                    source_row = source_matches[local_idx]
                    global_idx = source_mapping[local_idx]

                    if not target_matches_ref.empty:
                        closest_idx = self.find_closest_row(source_row, target_matches_ref, threshold=0.15)
                        target_row = target_matches_ref.iloc[closest_idx].to_dict()

                        changes = get_changes(source_row, target_row)
                        if changes:
                            differences[global_idx] = changes

                        # Remove the used target match to prevent reuse
                        target_matches_ref.drop(index=target_matches_ref.index[closest_idx], inplace=True)
                    else:
                        differences[global_idx] = None

            else:
                # Single source match
                source_row = source_matches[0]
                global_idx = source_mapping[0]

                if target_matches_ref is None or target_matches_ref.empty:
                    differences[global_idx] = None
                elif len(target_matches_ref) > 1:
                    closest_idx = self.find_closest_row(source_row, target_matches_ref, threshold=0.15)
                    target_row = target_matches_ref.iloc[closest_idx].to_dict()
                    changes = get_changes(source_row, target_row)
                    if changes:
                        differences[global_idx] = changes
                else:
                    # Single target match
                    target_row = target_matches_ref.iloc[0].to_dict()
                    changes = get_changes(source_row, target_row)
                    if changes:
                        differences[global_idx] = changes

        return differences

    @staticmethod
    def row_follows_condition(row, conditions: dict) -> bool:
        """Return True if all condition key-value pairs match the row."""
        return all(row.get(k) == v for k, v in conditions.items())

    @staticmethod
    def values_equal(a, b):
        """
        Compare two values for equality, accounting for types like NaN, datetime, and strings.
        """
        if pd.isnull(a) and pd.isnull(b):
            return True
        if isinstance(a, (float, int)) and isinstance(b, str):
            try:
                return a == float(b)
            except ValueError:
                return False
        if isinstance(a, str) and isinstance(b, (float, int)):
            try:
                return float(a) == b
            except ValueError:
                return False
        if isinstance(a, pd.Timestamp):
            try:
                b_dt = pd.to_datetime(b)
                return a == b_dt
            except Exception:
                return False
        if isinstance(b, pd.Timestamp):
            try:
                a_dt = pd.to_datetime(a)
                return b == a_dt
            except Exception:
                return False
        return a == b

    @staticmethod
    def find_closest_row(base: dict, df: pd.DataFrame, threshold: float = 0.75) -> Union[int, None]:
        """
        Return the row in the DataFrame that most closely matches the given base dictionary.

        Parameters:
        - base: Dictionary of column-value pairs to compare against.
        - df: Target DataFrame to search.
        - threshold: Minimum match ratio required.
        """
        if df.empty:
            return None

        def num_matches(row: pd.Series) -> int:
            return sum(base.get(key) == row.get(key) for key in base)

        match_counts = {idx: num_matches(row) for idx, row in df.iterrows()}

        if not match_counts:
            return None

        best_idx = max(match_counts, key=match_counts.get)
        max_matches = match_counts[best_idx]
        total_possible = len(base)

        if total_possible == 0 or (max_matches / total_possible) < threshold:
            return None

        # return df.loc[best_idx]
        return best_idx
    
    @staticmethod
    def valid_assign(value):
        return isinstance(value, (str, int, float, bool, datetime, date, time, Decimal)) or value is None
    
    @staticmethod
    def format_df(df: Union[pd.DataFrame, pd.Series]) -> Union[pd.DataFrame, pd.Series]:
        """
        Replace NaT and NaN with None, and convert datetime columns/values to native Python datetime.

        Parameters:
        - df: A pandas DataFrame or Series to format.
        """
        df = df.replace({pd.NaT: None, np.nan: None})

        if isinstance(df, pd.Series):
            # Convert datetime Series to Python datetimes
            if pd.api.types.is_datetime64_any_dtype(df):
                df = df.apply(lambda x: x.to_pydatetime() if pd.notnull(x) else None)
            return df
        
        for col in df.select_dtypes(include=["datetime64[ns]"]):
            df[col] = df[col].apply(lambda x: x.to_pydatetime() if pd.notnull(x) else None)
        return df