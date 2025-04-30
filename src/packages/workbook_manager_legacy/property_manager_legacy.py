import openpyxl
import os

from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from typing import Literal

# openpyxl cell properties: Value, Font, Fill (Background Color), Border, Alignment, Number Format, Comment, Hyperlink, Merge Cells, Unmerge Cells
CELL_PROPS = ["font", "fill", "border", "comment"]
# openpyxl sheet properties: Title, col/row height/width, lock row/col, tab color, hide/Unhide col/rows, lock cells, modify print settings, hide sheets, add header/footer, Set gridline visibility
SHEET_PROPS = ["sheet_state"]


class PropertyManager:
    property_store = {}
    hidden_sheets = []
    
    def __init__(self, read_file_path: str = None, sheet_names: list[str] = None) -> None:
        if read_file_path:
            if not os.path.exists(read_file_path):
                raise FileNotFoundError(f"Invalid file path provided to CommentManager: {read_file_path}")
            
            # Load the workbook
            wb = openpyxl.load_workbook(read_file_path)
            self.property_store = {name: {} for name in wb.sheetnames}
            self.hidden_sheets = [sheet_name for sheet_name in wb.sheetnames if wb[sheet_name].sheet_state == "hidden"]
            
        elif sheet_names:
            self.property_store = {name: {} for name in sheet_names}
            
    def append_comment(self, sheet: str, row: int, col: int, type: Literal["notice", "warning", "error"], comment: str):
        """Alter cell properties to indicate warning state."""
        
        if sheet not in self.property_store.keys():
            raise ValueError(f"The sheet '{sheet}' does not exist in the workbook.")
        
        if (row, col) not in self.property_store[sheet]:
            self.property_store[sheet][(row, col)] = [(type, comment)]
        else:
            self.property_store[sheet][(row, col)].append((type, comment))
        
    def apply_changes(self, write_file_path: str):
        """Apply cell property changes to an Excel file."""
        
        if not self.property_store:
            print("No cell properties to modify. Exiting Comment Manager.")
            return
        
        workbook = openpyxl.load_workbook(write_file_path)
        for sheet_name, affected_cells in self.property_store.items():
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"The sheet '{sheet_name}' does not exist in the workbook.")
            
            sheet_content = workbook[sheet_name]
            if sheet_name in self.hidden_sheets:
                sheet_content.sheet_state = "hidden"
            for (row, col), comments in affected_cells.items():
                cell = sheet_content.cell(row + 1, col + 1) # Plus 1 accounts for rows/cols being 1-based index
                comment_text = ""
                type_count = {"error": 0, "warning": 0, "notice": 0}
                for type, text in comments:                        
                    comment_text += f"{type.capitalize()}#{type_count[type]+1}: {text}\n"
                    type_count[type] += 1
                
                comment_obj = Comment(
                    comment_text, # Text
                    "Developer", # Author
                    150,  # Height in pixels
                    300   # Width in pixels
                )
                cell.comment = comment_obj

                # Set cell fill color based on the most severe type
                if type_count["error"] > 0:
                    fill_color = "bb2124"  # Red for errors
                elif type_count["warning"] > 0:
                    fill_color = "ffcc00"  # Yellow for warnings
                else:
                    fill_color = "5bc0de"  # Blue for notices
                          
                cell.fill = PatternFill(
                    start_color=fill_color,
                    end_color=fill_color,
                    fill_type='solid'
                )
        # Save the workbook with the modified properties
        workbook.save(write_file_path)