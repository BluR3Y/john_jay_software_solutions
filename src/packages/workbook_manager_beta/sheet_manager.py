import pandas as pd
import numpy as np
import logging
from typing import Union, Literal
from openpyxl import worksheet
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from openpyxl.cell import cell


class SheetManager:
    # openpyxl sheet properties: Title, col/row height/width, lock row/col, tab color, hide/Unhide col/rows, lock cells, modify print settings, hide sheets, add header/footer, Set gridline visibility
    assignable_sheet_props = ["sheet_state","column_dimensions"]
    # Reference
    # assigned_sheet_props = {
    #     "sheet_state": "visible"
    # }

    # openpyxl cell properties: Value, Font, Fill (Background Color), Border, Alignment, Number Format, Comment, Hyperlink, Merge Cells, Unmerge Cells
    assignable_cell_props = ["comment","fill","border","font","hyperlink","value","alignment","number_format"]
    # Reference
    # assigned_cell_props = {
    #     (1,5): {
    #         "comment": Comment("Test Comment Browski", "Developer", 150, 300),
    #         "fill": PatternFill(start_color='bb2124',end_color='bb2124', fill_type='solid')
    #     }
    # }

    issue_types = {
        "notice": {
            "priority": 2,
            "color": "5bc0de"
        },
        "warning": {
            "priority": 1,
            "color": "ffcc00"
        },
        "error": {
            "priority": 0,
            "color": "bb2124"
        }
    }

    def __init__(self, sheet_data: Union[pd.DataFrame, list[str]]):
        self.df = pd.DataFrame(columns=sheet_data) if isinstance(sheet_data, list) else sheet_data
        self.assigned_sheet_props = {}
        self.assigned_cell_props = {}

    def _append_data(self, dest_sheet: worksheet):
        for attr_name, attr_val in self.assigned_sheet_props.items():
            setattr(dest_sheet, attr_name, attr_val)

        for r_idx, row in self.df.iterrows():
            for c_idx, val in enumerate(row):
                curr_cell = dest_sheet.cell(row=r_idx + 1, column=c_idx + 1)

                if isinstance(curr_cell, cell.MergedCell):
                    continue

                curr_cell.value = val

                cell_attrs = self.assigned_cell_props.get((r_idx, c_idx))
                if cell_attrs:
                    for attr_name, attr_val in cell_attrs.items():
                        setattr(curr_cell, attr_name, attr_val)

    def series_indices(self, obj: pd.Series) -> list[int]:
        """
        Find row indices where the Series exactly matches a row in the sheet.
        """
        if not isinstance(obj, pd.Series):
            raise TypeError("obj must be a pandas Series")
        
        # Ensure obj has the same columns as the DataFrame for valid comparison
        if not all(col in self.df.columns for col in obj.index):
            raise ValueError("Series contains columns that do not exist in the sheet.")
        
        matching_indices = self.df.index[self.df.apply(lambda row: row.equals(obj), axis=1)]
        return matching_indices.tolist()
    
    def update_cell(self, row: Union[int, pd.Series], properties: dict) -> pd.Series:
        """
        Update values in a given row of a sheet and log the changes.

        Parameters:
        - row: Target row (by index or Series match).
        - properties: Dictionary of updates to apply.
        """
        # Determine the row index
        row_index = None
        if isinstance(row, pd.Series):
            matches = self.series_indices(self.df, row)
            if len(matches) != 1:
                raise ValueError("Series did not uniquely match a single row.")
            row_index = matches[0]
        elif isinstance(row, int):
            if not (0 <= row < self.df.shape[0]):
                raise ValueError("Row index out of bounds")
            row_index = row
        else:
            raise TypeError("Row must be an int or pandas Series")
        
        current = self.df.iloc[row_index]
        changing = {
            key: val for key, val in properties.items()
            if key in current.index and current[key] != val
        }

        # Warn on invalid property keys
        unknown_keys = set(properties.keys()) - set(current.index)
        if unknown_keys:
            logging.warning(f"Ignored unknown column(s): {unknown_keys}")
        
        if not changing:
            return current
        
        # Apply changes
        for key, val in changing.items():
            self.df.at[row_index, key] = val

        return self.df.iloc[row_index]
    
    def append_row(self, props: dict) -> pd.Series:
        """
        Append a new row to the sheet.

        Parameters:
        - props: Dictionary of column values.
        """
        # Create a new DataFrame
        new_row = pd.DataFrame({key: [value] for key, value in props.items()})
        # Append using pd.concat
        self.df = pd.concat([self.df, new_row], ignore_index=True)
        return new_row

    def delete_row(self, row: Union[int, pd.Series]) -> pd.Series:
        """
        Deletes a row from the specified sheet in the workbook.

        Parameters:
        - row (Union[int, pd.Series]): Row index or row content to identify the row to delete.
        """
        # Determine the row index
        row_index = None
        if isinstance(row, pd.Series):
            matches = self.series_indices(self.df, row)
            if len(matches) != 1:
                raise ValueError("Series did not uniquely match a single row.")
            row_index = matches[0]
        elif isinstance(row, int):
            if not (0 <= row < self.df.shape[0]):
                raise ValueError("Row index out of bounds")
            row_index = row
        else:
            raise TypeError("Row must be an int or pandas Series")
        
        # Capture data before deletion
        row_data = self.df.iloc[row_index]

        # Drop the row and reset index
        self.df = self.df.drop(index=row_index).reset_index(drop=True)
        return row_data
    
    def find(self, conditions: dict, return_one: bool = False, to_dict: str = None):
        """
        Retrieve rows matching specific conditions in a sheet.

        Parameters:
        - conditions: Column-value pairs to match.
        - return_one: Return only the first match if True.
        - to_dict: If specified, return result as a dict with given orientation.
        """
        # Ensure that conditions refer to existing columns
        invalid_columns = [col for col in conditions if col not in self.df.columns]
        if invalid_columns:
            raise KeyError(f"Invalid column names in conditions: {invalid_columns}")
        
        # Apply filtering
        mask = pd.Series(True, index=self.df.index)
        for column, value in conditions.items():
            mask &= (self.df[column] == value)
        
        filtered_rows = self.df[mask]
        
        if filtered_rows.empty:
            return filtered_rows
        
        formatted = self.format_df(filtered_rows)
        
        if to_dict:
            results = formatted.to_dict(orient=to_dict)
            return results[0] if return_one else results
        return formatted.iloc[0] if return_one else formatted
    
    def add_issue(self, row: int, col: int, type: Literal["notice", "warning", "error"], message: str):
        """Alter cell properties to indicate warning state."""
        num_rows, num_cols = self.df.shape
        if not (0 <= row < num_rows) or not (0 <= col < num_cols):
            raise ValueError("Index is out of bounds.")
        
        issue_props = self.issue_types[type]
        cell_key = (row, col)
        cell_props = self.assigned_cell_props.get(cell_key, {})
        comment_obj = cell_props.get("comment")

        # Initial comment and fill
        if not comment_obj:
            new_comment = Comment(f"Issues:\n* {type.capitalize()}#1 - {message}", "Developer", width=150, height=300)
            self.assigned_cell_props[cell_key] = {
                **cell_props,
                "comment": new_comment,
                "fill": PatternFill(start_color=issue_props["color"], end_color=issue_props["color"], fill_type='solid')
            }
            return

        # Extract and update issues from existing comment
        cell_messages = comment_obj.text.splitlines()
        try:
            issues_start_index = cell_messages.index("Issues:")
        except ValueError:
            issues_start_index = len(cell_messages)
            cell_messages.append("Issues:")

        issues_lines = cell_messages[issues_start_index + 1:]
        issues_by_type = {}
        min_priority = issue_props
        for issue in issues_lines:
            if issue.startswith("* "):
                try:
                    prefix, msg = issue[2:].split(" - ", 1)
                    issue_type = prefix.split("#")[0].lower()
                    issues_by_type.setdefault(issue_type, []).append(msg)
                    if self.issue_types[issue_type]["priority"] < min_priority["priority"]:
                        min_priority = self.issue_types[issue_type]
                except Exception:
                    continue

        # Append the new issue
        issues_by_type.setdefault(type, []).append(message)

        # Rebuild comment text
        formatted_issues = [
            f"* {itype.capitalize()}#{idx+1} - {msg}"
            for itype, messages in issues_by_type.items()
            for idx, msg in enumerate(messages)
        ]
        new_comment_text = "\n".join([*cell_messages[:issues_start_index + 1], *formatted_issues])
        self.assigned_cell_props[cell_key]["comment"].text = new_comment_text

        # Update fill based on highest priority
        self.assigned_cell_props[cell_key]["fill"] = PatternFill(
            start_color=min_priority['color'],
            end_color=min_priority['color'],
            fill_type='solid'
        )
        

        
    
    @staticmethod
    def format_df(df: Union[pd.DataFrame, pd.Series]) -> Union[pd.DataFrame, pd.Series]:
        """
        Replace NaT and NaN with None, and convert datetime columns/values to native Python datetime.

        Parameters:
        - df: A pandas DataFrame or Series to format.
        """
        df = df.replace({pd.NaT: None, np.nan: None})

        if isinstance(df, pd.Series):
            # Convert datetime Series to Python datetimes
            if pd.api.types.is_datetime64_any_dtype(df):
                df = df.apply(lambda x: x.to_pydatetime() if pd.notnull(x) else None)
            return df
        
        for col in df.select_dtypes(include=["datetime64[ns]"]):
            df[col] = df[col].apply(lambda x: x.to_pydatetime() if pd.notnull(x) else None)
        return df
    
    @staticmethod
    def values_equal(a, b):
        """
        Compare two values for equality, accounting for types like NaN, datetime, and strings.
        """
        if pd.isnull(a) and pd.isnull(b):
            return True
        if isinstance(a, (float, int)) and isinstance(b, str):
            try:
                return a == float(b)
            except ValueError:
                return False
        if isinstance(a, str) and isinstance(b, (float, int)):
            try:
                return float(a) == b
            except ValueError:
                return False
        if isinstance(a, pd.Timestamp):
            try:
                b_dt = pd.to_datetime(b)
                return a == b_dt
            except Exception:
                return False
        if isinstance(b, pd.Timestamp):
            try:
                a_dt = pd.to_datetime(a)
                return b == a_dt
            except Exception:
                return False
        return a == b

    @staticmethod
    def find_closest_row(base: dict, df: pd.DataFrame, threshold: float = 0.75) -> Union[pd.Series, None]:
        """
        Return the row in the DataFrame that most closely matches the given base dictionary.

        Parameters:
        - base: Dictionary of column-value pairs to compare against.
        - df: Target DataFrame to search.
        - threshold: Minimum match ratio required.
        """
        if df.empty:
            return None

        def num_matches(row: pd.Series) -> int:
            return sum(base.get(key) == row.get(key) for key in base)

        match_counts = {idx: num_matches(row) for idx, row in df.iterrows()}

        if not match_counts:
            return None

        best_idx = max(match_counts, key=match_counts.get)
        max_matches = match_counts[best_idx]
        total_possible = len(base)

        if total_possible == 0 or (max_matches / total_possible) < threshold:
            print(max_matches, total_possible, max_matches/total_possible)
            return None

        return df.loc[best_idx]