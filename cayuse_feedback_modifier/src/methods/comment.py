import openpyxl

# Add cell comments to the object's comment cache
def append_comment(self, sheet, row, col, comment):
    if sheet in self.sheets:
        if sheet not in self.comment_cache:
            self.comment_cache[sheet] = { f"{row}:{col}": comment }
        else:
            self.comment_cache[sheet].update({ f"{row}:{col}": comment })
    else:
        raise Exception(f"The sheet '{sheet}' does not exist in the workbook")

# Create the comments in the excel file that are stored in the object's cache
def create_comments(self, file_path):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    for sheet_name in self.comment_cache:
        if sheet_name in workbook.sheetnames:
            sheet_content = workbook[sheet_name]
            for cell_position in self.comment_cache[sheet_name]:
                row, col = cell_position.split(':')
                cell = sheet_content.cell(int(row) + 1, int(col) + 1)   # Plus 1 accounts for rows and columns being 1-based index
                comment = openpyxl.comments.Comment(self.comment_cache[sheet_name][cell_position], "Developer")
                cell.comment = comment
                comment.height = 150 # Height in pixels
                comment.width = 300 # Width in pixels
        else:
            raise Exception(f"The sheet '{sheet_name}' does not exist in the workbook")
    # Save the workbook with the comments
    workbook.save(file_path)